from django.shortcuts import render


from django.conf import settings

from .models import MCQ

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
import re
import pke
import string
from nltk.corpus import stopwords
from nltk.tokenize import sent_tokenize
from flashtext import KeywordProcessor
import requests
import random
from pywsd.similarity import max_similarity
from pywsd.lesk import adapted_lesk
from pywsd.lesk import simple_lesk
from pywsd.lesk import cosine_lesk
from nltk.corpus import wordnet as wn
from transformers import pipeline

from summarizer import Summarizer
from transformers import pipeline
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, redirect
from .forms import *
from .models import *
from django.http import HttpResponse

from django.shortcuts import render
import pke
import string
from nltk.corpus import stopwords

from fpdf import FPDF


import nltk
nltk.download('stopwords')
nltk.download('popular')
nltk.download('averaged_perceptron_tagger')
nltk.download('punkt')

from django.contrib.sessions.models import Session
from django.contrib.sessions.backends.db import SessionStore

@login_required
def generate_mcqs(request):
    generated_mcqs = []
    try:
        # Load all MCQs stored in the session
        mcqs_table_data = request.session.get('mcqs_table_data', [])
        if request.method == 'POST':
            input_text = request.POST.get('input_text', '')
            
           
            # Preprocess the input text to remove stopwords and punctuation
            stoplist = set(stopwords.words('english'))
            translator = str.maketrans('', '', string.punctuation)
            preprocessed_tokens = [word for word in input_text.split() if word.lower() not in stoplist]
            preprocessed_text = ' '.join(preprocessed_tokens)
            preprocessed_text = preprocessed_text.translate(translator)
            generated_mcqs = [] 
            # Initialize TextRank and load the preprocessed document
            extractor = pke.unsupervised.TextRank()
            extractor.load_document(input=preprocessed_text, language='en')

            # Perform candidate selection
            extractor.candidate_selection()

            summarizer = Summarizer()
            summarized_result = summarizer(input_text, min_length=60, max_length=1000, ratio=0.9)
            summarized_text = ''.join(summarized_result)
            # ... Other code for generating summarized_text ...
            extractor = pke.unsupervised.TopicRank()
            extractor.load_document(input=preprocessed_text)
            extractor.candidate_selection()

            # Get the keywords
            keywords = extractor.get_n_best(n=10)

            # Generate and filter keywords using PKE
            keywords = get_nouns_multipartite(input_text)
            filtered_keys = []
            for keyword in keywords:
                if keyword.lower() in summarized_text.lower():
                    filtered_keys.append(keyword)

            # Tokenize sentences and get sentence mapping for keywords
            sentences = tokenize_sentences(summarized_text)
            keyword_sentence_mapping = get_sentences_for_keyword(filtered_keys, sentences)

            # Generate keyword-distractor mapping using WordNet and ConceptNet
            key_distractor_list = {}
            for keyword in keyword_sentence_mapping:
                wordsense = get_wordsense(keyword_sentence_mapping[keyword][0], keyword)
                
                distractors = get_distractors_wordnet(wordsense, keyword)
                if len(distractors) == 0:
                    distractors = get_distractors_conceptnet(keyword)
                if len(distractors) != 0:
                    key_distractor_list[keyword] = distractors
                
                distractors = get_distractors_conceptnet(keyword)
                if len(distractors) != 0:
                    key_distractor_list[keyword] = distractors

            generated_mcqs = generate_mcqs_from_keyword_distractors(key_distractor_list, keyword_sentence_mapping)

            mcqs_table_data = []
            for index, mcq_text in enumerate(generated_mcqs, start=1):
                question = mcq_text.split('\n')[0]
                options_start_index = mcq_text.index('a)')  # Assuming options start with 'a)'
                options_end_index = mcq_text.index('More options:')
                options_text = mcq_text[options_start_index:options_end_index]
                options = [option.strip() for option in options_text.split('\n') if option.strip()]
                correct_answer = mcq_text.split('Correct answer:')[1].strip()
                mcqs_table_data.append({
                    'question': question,
                    'options': options,
                    'correct_answer': correct_answer
                })

                # Load previously generated MCQs if available in the session
            session_key = request.session.session_key
            if not session_key:
                request.session.save()
            previous_mcqs = request.session.get('mcqs_table_data', [])
            mcqs_table_data = previous_mcqs + mcqs_table_data
            # try:
            #     session = Session.objects.get(session_key=session_key)
            #     session_data = session.get_decoded()
            #     if 'mcqs_table_data' in session_data:
            #         previous_mcqs = session_data['mcqs_table_data']
            #         mcqs_table_data += previous_mcqs
            # except Session.DoesNotExist:
            #     # Handle the case where the session does not exist
            #     pass

            # Combine the newly generated MCQs with the previous MCQs
            #mcqs_table_data += request.session.get('mcqs_table_data', [])
            # Save the updated MCQs in the session
            request.session['mcqs_table_data'] = mcqs_table_data

            print('hello')
            return render(request, 'dashboard/generate_mcqs.html', {'input_text': input_text, 'generated_mcqs': mcqs_table_data})
        
        # If the request method is not POST (e.g., when initially loading the page)
        # Retrieve MCQs stored in the session
        mcqs_table_data = request.session.get('mcqs_table_data', [])
        print('yooooo')
        return render(request, 'dashboard/generate_mcqs.html', { 'generated_mcqs': mcqs_table_data})
    
    except IndexError as e:
        error_message = f"An error occurred: {str(e)}"  
        return render(request, 'dashboard/generate_mcqs.html', { 'error_message': error_message})
    except Exception as e:
        error_message = f"An error occurred: {str(e)}"
        return render(request, 'dashboard/generate_mcqs.html', { 'error_message': error_message})
    return render(request, 'dashboard/generate_mcqs.html')

#----------------------------------------------------------

#What is SQL? SQL stands for Structured Query Language. SQL lets you access and manipulate databases. SQL became a standard of the American National Standards Institute (ANSI) in 1986, and of the International Organization for Standardization (ISO) in 1987.


# PKE code for generating keywords
def get_nouns_multipartite(text):
    out = []

    extractor = pke.unsupervised.MultipartiteRank()
    extractor.load_document(input=text)
    #    not contain punctuation marks or stopwords as candidates.
    pos = {'PROPN'}
    #pos = {'VERB', 'ADJ', 'NOUN'}
    stoplist = list(string.punctuation)
    stoplist += ['-lrb-', '-rrb-', '-lcb-', '-rcb-', '-lsb-', '-rsb-']
    stoplist += stopwords.words('english')
    extractor.candidate_selection(pos=pos)
    # 4. build the Multipartite graph and rank candidates using random walk,
    #    alpha controls the weight adjustment mechanism, see TopicRank for
    #    threshold/method parameters.
    extractor.candidate_weighting(alpha=1.1,
                                  threshold=0.75,
                                  method='average')
    keyphrases = extractor.get_n_best(n=20)

    for key in keyphrases:
        out.append(key[0])

    return out
    # ... Existing code for PKE keyword extraction ...

# Tokenize sentences
def tokenize_sentences(text):
    sentences = sent_tokenize(text)
    sentences = [sentence.strip() for sentence in sentences if len(sentence) > 20]
    return sentences

# Get sentence mapping for keywords
def get_sentences_for_keyword(keywords, sentences):
    keyword_processor = KeywordProcessor()
    keyword_sentences = {}
    for word in keywords:
        keyword_sentences[word] = []
        keyword_processor.add_keyword(word)
    for sentence in sentences:
        keywords_found = keyword_processor.extract_keywords(sentence)
        for key in keywords_found:
            keyword_sentences[key].append(sentence)
    for key in keyword_sentences.keys():
        values = keyword_sentences[key]
        values = sorted(values, key=len, reverse=True)
        keyword_sentences[key] = values
    return keyword_sentences
    
    # ... Existing code for keyword sentence mapping ...

# Distractors from Wordnet
# views.py
from nltk.corpus import wordnet as wn

def get_distractors_wordnet(syn, word):
    distractors = []
    word = word.lower()
    orig_word = word
    if len(word.split()) > 0:
        word = word.replace(" ", "_")
    if syn is None:
        print(f"Wordsense for '{word}' is None")
        return distractors
    hypernym = syn.hypernyms()
    if len(hypernym) == 0:
        return distractors
    for item in hypernym[0].hyponyms():
        name = item.lemmas()[0].name()
        if name == orig_word:
            continue
        name = name.replace("_", " ")
        name = " ".join(w.capitalize() for w in name.split())
        if name is not None and name not in distractors:
            distractors.append(name)
    return distractors


    # ... Existing code for WordNet distractors ...

# Get wordsense using PyWSD
from nltk.corpus import wordnet as wn

def get_wordsense(sent, word):
    word = word.lower()
    
    if len(word.split()) > 0:
        word = word.replace(" ", "_")
    
    synsets = wn.synsets(word, 'n')
    if synsets:
        wup = max_similarity(sent, word, 'wup', pos='n')
        lesk_output = simple_lesk(sent, word, pos='n')
        if wup in synsets and lesk_output in synsets:
            lowest_index = min(synsets.index(wup), synsets.index(lesk_output))
            return synsets[lowest_index]
        elif wup in synsets:
            return wup
        elif lesk_output in synsets:
            return lesk_output
    return None
    # ... Existing code for getting wordsense ...

# Distractors from ConceptNet
def get_distractors_conceptnet(word):
    word = word.lower()
    original_word = word
    if len(word.split()) > 0:
        word = word.replace(" ", "_")
    distractor_list = []
    url = f"http://api.conceptnet.io/query?node=/c/en/{word}/n&rel=/r/PartOf&start=/c/en/{word}&limit=5"
    obj = requests.get(url).json()

    for edge in obj['edges']:
        link = edge['end']['term']
        url2 = f"http://api.conceptnet.io/query?node={link}&rel=/r/PartOf&end={link}&limit=10"
        obj2 = requests.get(url2).json()
        for edge in obj2['edges']:
            word2 = edge['start']['label']
            if word2 not in distractor_list and original_word.lower() not in word2.lower():
                distractor_list.append(word2)

    return distractor_list
    # ... Existing code for ConceptNet distractors ...

# Generate MCQs using keyword-distractor mapping
def generate_mcqs_from_keyword_distractors(key_distractor_list, keyword_sentence_mapping):
    mcqs = []

    index = 1
    for each in key_distractor_list:
        sentence = keyword_sentence_mapping[each][0]
        pattern = re.compile(each, re.IGNORECASE)
        output = pattern.sub(" _______ ", sentence)
        question = f"{index}) {output}\n"

        correct_answer = each.capitalize()
        choices = [correct_answer] + key_distractor_list[each]
        #choices = [each.capitalize()] + key_distractor_list[each]
        top4choices = choices[:4]
        random.shuffle(top4choices)
        optionchoices = ['a', 'b', 'c', 'd']
        for idx, choice in enumerate(top4choices):
            question += f"\t\n{optionchoices[idx]}) {choice}\n"
        question += f"\nMore options: {choices[4:20]}\n\n"
        question += f"\nCorrect answer: {correct_answer}\n\n"
        mcqs.append(question)
        index = index + 1

    return mcqs

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
from django.shortcuts import redirect

def remove_mcq(request):
    if request.method == 'POST':
        mcq_index = int(request.POST.get('mcq_index', -1))
        
        if mcq_index >= 0:
            # Load the existing MCQs from the session
            mcqs_table_data = request.session.get('mcqs_table_data', [])

            # Check if the specified MCQ index is valid
            if mcq_index < len(mcqs_table_data):
                # Remove the MCQ from the list
                removed_mcq = mcqs_table_data.pop(mcq_index)

                # Update the session with the modified MCQ list
                request.session['mcqs_table_data'] = mcqs_table_data

                # Optionally, you can add a message to indicate successful removal
                request.session['message'] = f"MCQ '{removed_mcq['question']}' removed successfully."

    return redirect('dashboard:generate_mcqs')





from django.shortcuts import redirect

def remove_question(request):
    if request.method == 'POST':
        mcq_index = int(request.POST.get('mcq_index', -1))
        
        if mcq_index >= 0:
            # Load the existing MCQs from the session
            generated_questions = request.session.get('generated_questions', [])

            # Check if the specified MCQ index is valid
            if mcq_index < len(generated_questions):
                # Remove the MCQ from the list
                removed_mcq = generated_questions.pop(mcq_index)

                # Update the session with the modified MCQ list
                request.session['generated_questions'] = generated_questions

                # Optionally, you can add a message to indicate successful removal
                #request.session['message'] = f"MCQ '{removed_mcq['question']}' removed successfully."

    return redirect('dashboard:theorytopdf')


#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# New view for generating theoretical questions
import openai
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse


openai.api_key = "    Api Key   "

@login_required
def generate_questions(request):
    input_text = ""
    new_questions = []
    questions = request.session.get('generated_questions', [])

    if request.method == 'POST':
        input_text = request.POST.get('input_text', '')

        if input_text:
            new_questions = generate_questions_from_text(input_text)
            questions += new_questions
            request.session['generated_questions'] = questions

    return render(request, 'dashboard/generate_questions.html', {'questions': questions, 'new_questions': new_questions, 'input_text': input_text})


def generate_questions_from_text(input_text):
    try:
        model_engine = "text-davinci-002"  # choose a language model to use

        # Generate questions using GPT-3
        prompt = f"Generate questions based on the following text:\n\n{input_text}\n\n"
        response = openai.Completion.create(
            engine=model_engine,
            prompt=prompt,
            max_tokens=1024,
            n=2,  # You can adjust the number of questions generated
            stop=None,
            temperature=0.5,
        )
        generated_text = response.choices[0].text.strip()

        # Split the generated text into questions
        questions = generated_text.split("\n")

        return questions
    except Exception as e:
        error_message = f"An error occurred: {str(e)}"
        return [error_message]




# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Convert MCQ from PDF OR TEXT FILE
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from transformers import pipeline
import string
from nltk.corpus import stopwords
import pke
from summarizer import Summarizer


@login_required
def mcqtopdf(request):

    generated_mcqs = []  # Initialize the variable
    form_submitted = False
    error_message = None  # Initialize the error message to None

    try:
        # Load all MCQs stored in the session
        mcqsform_table_data = request.session.get('mcqs_table_data', [])

        if request.method == 'POST':
            form_submitted = True
            uploaded_file = request.FILES.get('input_file')
            if uploaded_file:
                fs = FileSystemStorage()
                filename = fs.save(uploaded_file.name, uploaded_file)
                filepath = fs.url(filename)
                with fs.open(filename) as file:
                    input_text = file.read().decode('utf-8')

                stoplist = set(stopwords.words('english'))
                translator = str.maketrans('', '', string.punctuation)
                preprocessed_tokens = [word for word in input_text.split() if word.lower() not in stoplist]
                preprocessed_text = ' '.join(preprocessed_tokens)
                preprocessed_text = preprocessed_text.translate(translator)

                # ... Your existing code to process the input_text ...

                generated_mcqs = [] 
                # Initialize TextRank and load the preprocessed document
                extractor = pke.unsupervised.TextRank()
                extractor.load_document(input=preprocessed_text, language='en')

                # Perform candidate selection
                extractor.candidate_selection()

                summarizer = Summarizer()
                summarized_result = summarizer(input_text, min_length=60, max_length=500, ratio=0.4)
                summarized_text = ''.join(summarized_result)
                # ... Other code for generating summarized_text ...
                extractor = pke.unsupervised.TopicRank()
                extractor.load_document(input=preprocessed_text)
                extractor.candidate_selection()

                # Get the keywords
                keywords = extractor.get_n_best(n=10)

                # Generate and filter keywords using PKE
                keywords = get_nouns_multipartite(input_text)
                filtered_keys = []
                for keyword in keywords:
                    if keyword.lower() in summarized_text.lower():
                        filtered_keys.append(keyword)

                # Tokenize sentences and get sentence mapping for keywords
                sentences = tokenize_sentences(summarized_text)
                keyword_sentence_mapping = get_sentences_for_keyword(filtered_keys, sentences)

                # Generate keyword-distractor mapping using WordNet and ConceptNet
                key_distractor_list = {}
                for keyword in keyword_sentence_mapping:
                    wordsense = get_wordsense(keyword_sentence_mapping[keyword][0], keyword)
                
                    distractors = get_distractors_wordnet(wordsense, keyword)
                    if len(distractors) == 0:
                        distractors = get_distractors_conceptnet(keyword)
                    if len(distractors) != 0:
                        key_distractor_list[keyword] = distractors
                
                    distractors = get_distractors_conceptnet(keyword)
                    if len(distractors) != 0:
                        key_distractor_list[keyword] = distractors
                        
                generated_mcqs = generate_mcqs_from_keyword_distractors(key_distractor_list, keyword_sentence_mapping)
                mcqsform_table_data = []
                for index, mcq_text in enumerate(generated_mcqs, start=1):
                    question = mcq_text.split('\n')[0]
                    options_start_index = mcq_text.index('a)')  # Assuming options start with 'a)'
                    options_end_index = mcq_text.index('More options:')
                    options_text = mcq_text[options_start_index:options_end_index]
                    options = [option.strip() for option in options_text.split('\n') if option.strip()]
                    correct_answer = mcq_text.split('Correct answer:')[1].strip()
                    mcqsform_table_data.append({
                        'question': question,
                        'options': options,
                        'correct_answer': correct_answer
                    })

                previous_mcqs = request.session.get('mcqs_table_data', [])
                mcqsform_table_data = previous_mcqs + mcqsform_table_data

                # Append the newly generated MCQs to the existing ones
                #mcqsform_table_data += generated_mcqs

                # Save the updated MCQs in the session request.session['mcqs_table_data'] = mcqs_table_data
                request.session['mcqs_table_data'] = mcqsform_table_data

                
                return render(request, 'dashboard/mcqtopdf.html', {
                    'generated_mcqs': mcqsform_table_data,
                    'form_submitted': form_submitted,
                    'error_message': error_message
                })
        mcqsform_table_data = request.session.get('mcqs_table_data', [])
        
        return render(request, 'dashboard/mcqtopdf.html', {
            'generated_mcqs': mcqsform_table_data,
        })

    except Exception as e:
        error_message = f"An error occurred: {str(e)}"
        return render(request, 'dashboard/mcqtopdf.html', {
            'generated_mcqs': generated_mcqs,
            'form_submitted': form_submitted,
            'error_message': error_message,
        })




from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

# Your other imports and settings...

@login_required
def theorytopdf(request):
    input_text = ""
    new_questions = []
    questions = request.session.get('generated_questions', [])

    if request.method == 'POST':
        if 'input_file' in request.FILES:
            input_file = request.FILES['input_file']
            input_text = input_file.read().decode('utf-8')
        else:
            input_text = request.POST.get('input_text', '')

        if input_text:
            new_questions = generate_questions_from_text(input_text)
            questions += new_questions
            request.session['generated_questions'] = questions

    return render(request, 'dashboard/theorytopdf.html', {'questions': questions, 'new_questions': new_questions, 'input_text': input_text})


#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

from django.shortcuts import render

def view_profile(request):
    if request.user.is_authenticated:
        user = request.user
        context = {'user': user}
        return render(request, 'dashboard/viewprofile.html', context)
    

# Create your views here.
def index(request):
    return render(request, 'dashboard/index.html')




def tables(request):
    return render(request, "dashboard/tables.html")




def grid(request):
    return render(request, "dashboard/grid.html")




def form_basic(request):
    return render(request, "dashboard/form_basic.html")




def form_wizard(request):
    return render(request, "dashboard/form_wizard.html")




def buttons(request):
    return render(request, "dashboard/buttons.html")






def elements(request):
    return render(request, "dashboard/elements.html")





from .forms import *
from .models import *
from django.http import HttpResponse
def manual_mcq(request):
    if request.method == 'POST':
        print(request.POST)
        questions=QuesModel.objects.all()
        score=0
        wrong=0
        correct=0
        total=0
        for q in questions:
            total+=1
            print(request.POST.get(q.question))
            print(q.ans)
            print()
            if q.ans ==  request.POST.get(q.question):
                score+=10
                correct+=1
            else:
                wrong+=1
        percent = score/(total*10) *100
        context = {
            'score':score,
            'time': request.POST.get('timer'),
            'correct':correct,
            'wrong':wrong,
            'percent':percent,
            'total':total
        }
        return render(request,'dashboard/manual_result.html',context)
    else:
        questions=QuesModel.objects.all()
        context = {
            'questions':questions
        }
        return render(request, "dashboard/manual_mcq.html",context)


def manual_result(request):
    return render(request, "dashboard/manual_result.html")


def addquestion(request):
    if request.user.is_staff:
        form=addQuestionform()
        if(request.method=='POST'):
            form=addQuestionform(request.POST)
            if(form.is_valid()):
                form.save()
                return redirect('/')
        context={'form':form}
        return render(request,"dashboard/addquestion.html",context)
    else: 
        
        return render(request, "dashboard/addquestion.html")


from django.shortcuts import render
import pke
import string
from nltk.corpus import stopwords

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

from django.shortcuts import render
from django.http import HttpResponse
from fpdf import FPDF

def generate_pdf(request):
    pdf_output = None
    
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')
        
        # Create a PDF instance
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        # Add content to the PDF
        pdf.cell(200, 10, f"Name: {name}", ln=True)
        pdf.cell(200, 10, f"Email: {email}", ln=True)
        pdf.multi_cell(0, 10, f"Message:\n{message}")
        
        # Generate PDF content
        pdf_output = pdf.output(dest='S').encode('latin1').decode('latin1')
    
    return render(request, 'dashboard/generate_pdf.html', {'pdf_output': pdf_output})


#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Generate PDF

# views.py
from django.contrib.sessions.models import Session
from django.http import FileResponse
from django.shortcuts import render
from fpdf import FPDF
import os
import tempfile
def generate_question_paper(request):
    # Retrieve generated_mcqs from the session
    mcqs_table_data = request.session.get('mcqs_table_data', [])
    generated_questions = request.session.get('generated_questions', [])
    
    #generated_questions = request.session['mcqsform_table_data'] 
    # Your PDF generation code here
    # Create an instance of FPDF
    try:
        # Create a temporary file to store the PDF
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf_file:
            pdf = FPDF()
            pdf.add_page()
            
            # Set font for the PDF
            pdf.image("static/images/jaihindlogo.png",x=30, y=9,w=150)
            pdf.set_font(family="Arial", style="B", size=9)
            pdf.ln(28)
            pdf.cell(0, h=5, txt="March 2023", align="C", border="TLR", ln=1)
            pdf.cell(0, h=5, txt="Examination: End Semester -IV Examination (UG Programmes)", align="C", border="BLR", ln=1)
            pdf.cell(97.5, h=10, txt="Name of the Constituent College: Jai Hind College", border="LB")
            pdf.cell(92.5, h=10, txt="Name of the department: Computer Science", border="LBR", ln=1)
            pdf.cell(w=50,h=10,txt="Duration: 2 Hr",border="LB")
            pdf.cell(w=70,h=10,txt="Name of Course: Internet of Things",border="LB")
            pdf.cell(w=70,h=10,txt="Max. Marks: 60",border="LRB", ln=1)

            pdf.cell(0,h=3,txt="Instructions:",border="LR",ln=1)
            pdf.cell(0,h=3,txt="i) All Questions are Compulsory",border="LR",ln=1)
            pdf.cell(0,h=3,txt="ii)Draw diagrams Whereever necessary",border="LR",ln=1)
            pdf.cell(0,h=3,txt="iii)Figures to right indicate full marks.",border="LRB",ln=1)

            pdf.cell(0,h=5,border=False,ln=1)

            pdf.cell(0,h=5,txt="Questions",border="TLR",ln=1)
            # if mcqs_table_data:
            #     for mcq_dict in mcqs_table_data:
            #         # Extract relevant information from the dictionary
            #         question = mcq_dict.get('question', '')
            #         options = mcq_dict.get('options', [])
                    
            #         # Format question and options as a string
            #         mcq_text = f"{question}\n"
            #         mcq_text += "\n".join(options) + "\n"
                    
            #         # Add the formatted MCQ text to the PDF
            #         pdf.multi_cell(0, 10, mcq_text)
                    
            #         # Add space between MCQs
            #         pdf.ln(10)
            
            # pdf.cell(w=15,h=5,txt="Marks",border="TLR",align="C",ln=1)
            
            if mcqs_table_data:
                
                for mcq_dict in mcqs_table_data:
                    
                    question = mcq_dict.get('question', '')
                    options = mcq_dict.get('options', [])
                    
                    
                    mcq_text = f"{question}\n"
                    mcq_text += "\n".join(options) + "\n"
                    
                    pdf.multi_cell(0,10,txt=mcq_text,border=1)
                    
            
            
                  # Add space between MCQs
                    #pdf.ln(1)

            if generated_questions:
                #formatted_questions = []  # Create an empty list to store formatted questions
                for ques_dict in generated_questions:
                    
                    # Append the formatted question to the list
                    # pdf.multi_cell(0, 10, ques_dict)
                    pdf.multi_cell(0,h=5,txt=ques_dict,border="TLRB")
                    # Add space between questions
                            
            
            # pdf.cell(w=15,h=180,border="TRB")
            
            # Save the PDF to a temporary file
            

            pdf.output(tmp_pdf_file.name)
            
            # Return the PDF as a FileResponse
            with open(tmp_pdf_file.name, 'rb') as pdf_file:
                pdf_content = pdf_file.read()
                response = HttpResponse(pdf_content, content_type='application/pdf')
                #response = FileResponse(pdf_file, content_type='application/pdf')
                response['Content-Disposition'] = f'inline; filename="{os.path.basename(tmp_pdf_file.name)}"'
        
        return response
    except Exception as e:
        # Handle exceptions appropriately
        return render(request, 'dashboard/generate_mcqs.html', {'error_message': str(e)})
        
# views.py

def display_question_paper(request):
    return render(request, 'dashboard/generate_pdf.html')
  


# from django.shortcuts import render
# from django.http import HttpResponse
# from fpdf import FPDF
# import tempfile
# import os


from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.sessions.models import Session
from fpdf import FPDF
import os
import tempfile


def generate_question_paper_form(request):
    try:
        if request.method == 'POST':
            # Get user input from the form
            college_name = request.POST.get('college_name', '')
            department_name = request.POST.get('department_name', '')
            date = request.POST.get('date', '')
            examination = request.POST.get('examination', '')
            duration = request.POST.get('duration', '')
            course_name = request.POST.get('course_name', '')
            max_marks = request.POST.get('max_marks', '')

            # Retrieve generated_mcqs from the session
            mcqs_table_data = request.session.get('mcqs_table_data', [])

            generated_questions = request.session.get('generated_questions', [])

            # Create an instance of FPDF
            pdf = FPDF()
            pdf.add_page()

            # Set font for the PDF
            pdf.set_font(family="Arial", style="B", size=9)
            pdf.image("static/images/jaihindlogo.png",x=30, y=9,w=150)
            pdf.ln(28)
            pdf.cell(0, h=5, txt=f"{date}", align="C", border="TLR", ln=1)
            pdf.cell(0, h=5, txt=f"Examination: {examination}", align="C", border="BLR", ln=1)
            pdf.cell(97.5, h=10, txt=f"Name of the Constituent College: {college_name} ", border="LB")
            pdf.cell(92.5, h=10, txt=f"Name of the department: {department_name}", border="LBR", ln=1)
            pdf.cell(w=50,h=10,txt=f"Duration: {duration}",border="LB")
            pdf.cell(w=70,h=10,txt=f"Name of Course: {course_name}",border="LB")
            pdf.cell(w=70,h=10,txt=f"Max. Marks: {max_marks}",border="LRB", ln=1)

            pdf.cell(0,h=3,txt="Instructions:",border="LR",ln=1)
            pdf.cell(0,h=3,txt="i) All Questions are Compulsory",border="LR",ln=1)
            pdf.cell(0,h=3,txt="ii)Draw diagrams Wherever necessary",border="LR",ln=1)
            pdf.cell(0,h=3,txt="iii)Figures to right indicate full marks.",border="LRB",ln=1)

            pdf.cell(0,h=5,border=False,ln=1)

            pdf.cell(0,h=5,txt="Questions",border="TLR",ln=1)
            # Include MCQs if the checkbox is selected
            if 'mcqCheckbox' in request.POST and mcqs_table_data:
                for mcq_dict in mcqs_table_data:
                    
                    question = mcq_dict.get('question', '')
                    options = mcq_dict.get('options', [])
                    
                    
                    mcq_text = f"{question}\n"
                    mcq_text += "\n".join(options) + "\n"
                    
                    pdf.multi_cell(0,10,txt=mcq_text,border=1)

            # Include Questions if the checkbox is selected
            if 'questionsCheckbox' in request.POST and generated_questions:
                for ques_dict in generated_questions:
                    pdf.multi_cell(0, 5, txt=ques_dict, border="TLRB")

            # Save the PDF to a temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf_file:
                pdf.output(tmp_pdf_file.name)

            # Serve the PDF as an HttpResponse
            with open(tmp_pdf_file.name, 'rb') as pdf_file:
                pdf_content = pdf_file.read()
                response = HttpResponse(pdf_content, content_type='application/pdf')
                response['Content-Disposition'] = f'inline; filename="{os.path.basename(tmp_pdf_file.name)}"'
           
            return response

        return render(request, 'dashboard/generate_question_form.html')
    except Exception as e:
        # Handle exceptions appropriately
        return render(request, 'dashboard/generate_question_form.html', {'error_message': str(e)})



def display_question_form(request):
    return render(request, "dashboard/generate_question_form.html")

import openpyxl
from django.shortcuts import render
from django.contrib import messages
from django.http import HttpResponse

def upload_and_display_excel(request):
    try:
        data_from_excel = []  # Initialize an empty list to store Excel data

        if request.method == 'POST':
            # Handle the uploaded Excel file
            uploaded_excel_file = request.FILES.get('excel_file')

            if uploaded_excel_file:
                # Check if the uploaded file is an Excel file (xlsx or xls)
                if not uploaded_excel_file.name.lower().endswith(('.xlsx', '.xls')):
                    print("DEBUG: Please upload a valid Excel file.")
                    messages.error(request, 'Please upload a valid Excel file.')
                else:
                    try:
                        # Load the Excel file into openpyxl
                        excel_workbook = openpyxl.load_workbook(uploaded_excel_file, data_only=True)
                        excel_sheet = excel_workbook.active

                        # Iterate through rows and append data from the first column to the list
                        for row in excel_sheet.iter_rows(values_only=True):
                            data_from_excel.append(row[0])

                    except Exception as e:
                        print(f"Error processing Excel file: {str(e)}")
                        messages.error(request, f"Error processing Excel file: {str(e)}")

            else:
                print("DEBUG: No Excel file uploaded.")
                messages.error(request, 'Please upload an Excel file.')

        return render(request, 'dashboard/upload_excel.html', {'data_from_excel': data_from_excel})

    except Exception as e:
        # Handle exceptions appropriately
        return render(request, 'dashboard/upload_excel.html', {'error_message': str(e)})
